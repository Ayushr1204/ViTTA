"""
Excel (.xlsx) exporter for the final ViTTA trajectory dataset.

Produces a multi-sheet workbook:
  Sheet 1 — Trajectory Data  (wide format: one row per vehicle,
                               columns expand for each 1-sec timestamp)
  Sheet 2 — Track Summary    (one row per unique track with statistics)
  Sheet 3 — Metadata         (video info, processing config)
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from vitta.tracking.resampler import ResampledRecord
from vitta.class_names import class_name

logger = logging.getLogger(__name__)


# ── Styles ────────────────────────────────────────────────────────────

_HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
_HEADER_BORDER = Border(
    bottom=Side(style="thin", color="000000"),
    right=Side(style="thin", color="D9D9D9"),
)

_SUBHEADER_FONT = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
_SUBHEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

_DATA_FONT = Font(name="Calibri", size=10)
_DATA_ALIGNMENT = Alignment(horizontal="center", vertical="center")
_ALT_ROW_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

# Per-timestamp metrics that become column groups
_TIME_FIELDS = ["cx_px", "cy_px", "speed_px_per_sec", "cumulative_distance_px"]
_TIME_FIELD_LABELS = {
    "cx_px": "Centroid X (px)",
    "cy_px": "Centroid Y (px)",
    "speed_px_per_sec": "Speed (px/s)",
    "cumulative_distance_px": "Cumul. Distance (px)",
}


def _style_header_row(ws, row: int, num_cols: int, font=None, fill=None) -> None:
    """Apply styling to a header row."""
    font = font or _HEADER_FONT
    fill = fill or _HEADER_FILL
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = font
        cell.fill = fill
        cell.alignment = _HEADER_ALIGNMENT
        cell.border = _HEADER_BORDER


def _auto_column_width(ws, min_width: int = 10, max_width: int = 22) -> None:
    """Set column widths based on content."""
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        adjusted = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


def _apply_alternating_rows(ws, start_row: int, end_row: int, num_cols: int) -> None:
    """Apply subtle alternating-row shading for readability."""
    for row in range(start_row, end_row + 1):
        if row % 2 == 0:
            for col in range(1, num_cols + 1):
                ws.cell(row=row, column=col).fill = _ALT_ROW_FILL


# ═══════════════════════════════════════════════════════════════════════
# Public API
# ═══════════════════════════════════════════════════════════════════════

class ExcelExporter:
    """
    Export resampled track data to a multi-sheet Excel workbook.

    The trajectory sheet uses a **wide format**: each vehicle occupies
    exactly one row, and columns expand horizontally for each timestamp
    (cx, cy, speed, distance at t=0s, t=1s, t=2s, …).

    Usage::

        exporter = ExcelExporter()
        exporter.export(
            records=resampled_records,
            output_path="output/vitta_results.xlsx",
            metadata={"video": "traffic.mp4", ...},
        )
    """

    _FIXED_HEADERS = ["Track ID", "Class Name", "First Seen (s)", "Last Seen (s)"]

    _SUMMARY_HEADERS = [
        "Track ID",
        "Class ID",
        "Class Name",
        "First Seen (s)",
        "Last Seen (s)",
        "Duration (s)",
        "Total Distance (px)",
        "Avg Speed (px/s)",
        "Max Speed (px/s)",
        "Num Observations",
        "Avg Confidence",
    ]

    def export(
        self,
        records: List[ResampledRecord],
        output_path: str | Path,
        metadata: Optional[Dict[str, Any]] = None,
    ) -> Path:
        """
        Write the Excel workbook.

        Args:
            records:     List of ResampledRecord from TrackResampler.
            output_path: Destination .xlsx file path.
            metadata:    Optional dict of video/pipeline metadata.

        Returns:
            The resolved output Path.
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()

        # ── Sheet 1: Trajectory Data (wide format) ───────────────────
        ws_traj = wb.active
        ws_traj.title = "Trajectory Data"
        self._write_trajectory_sheet_wide(ws_traj, records)

        # ── Sheet 2: Track Summary ───────────────────────────────────
        ws_summary = wb.create_sheet("Track Summary")
        self._write_summary_sheet(ws_summary, records)

        # ── Sheet 3: Metadata ────────────────────────────────────────
        ws_meta = wb.create_sheet("Metadata")
        self._write_metadata_sheet(ws_meta, metadata or {})

        wb.save(str(output_path))
        num_tracks = len(set(r.track_id for r in records))
        logger.info(
            f"Excel workbook saved: {output_path} "
            f"({num_tracks} tracks, wide format)"
        )
        return output_path

    # ── Private sheet writers ─────────────────────────────────────────

    def _write_trajectory_sheet_wide(
        self,
        ws,
        records: List[ResampledRecord],
    ) -> None:
        """
        Write trajectory data in wide format.

        Layout:
          Row 1 (header):  Track ID | Class Name | First Seen | Last Seen |  t=0s          |  t=1s          | ...
          Row 2 (sub-hdr): (empty)  | (empty)    | (empty)    | (empty)   | CX|CY|Spd|Dist | CX|CY|Spd|Dist | ...
          Row 3+:          data rows (one per track)
        """
        if not records:
            ws.append(["No tracking data available."])
            return

        # ── Group records by track_id ────────────────────────────────
        tracks: Dict[int, List[ResampledRecord]] = {}
        for rec in records:
            tracks.setdefault(rec.track_id, []).append(rec)

        # ── Collect all unique timestamps across all tracks ──────────
        all_timestamps: set[float] = set()
        for recs in tracks.values():
            for r in recs:
                all_timestamps.add(r.timestamp_sec)
        sorted_timestamps = sorted(all_timestamps)

        num_fixed = len(self._FIXED_HEADERS)
        num_fields = len(_TIME_FIELDS)
        num_time_cols = len(sorted_timestamps) * num_fields
        total_cols = num_fixed + num_time_cols

        # ── Row 1: Main header (merged-style labels) ────────────────
        header_row_1 = list(self._FIXED_HEADERS)
        for ts in sorted_timestamps:
            ts_label = f"t = {ts:.0f}s"
            header_row_1.append(ts_label)
            # Pad remaining fields in this group with empty strings
            for _ in range(num_fields - 1):
                header_row_1.append("")
        ws.append(header_row_1)

        # Merge cells for each timestamp group in row 1
        for i, _ in enumerate(sorted_timestamps):
            start_col = num_fixed + 1 + i * num_fields
            end_col = start_col + num_fields - 1
            ws.merge_cells(
                start_row=1, start_column=start_col,
                end_row=1, end_column=end_col,
            )

        _style_header_row(ws, 1, total_cols)

        # ── Row 2: Sub-headers (field names under each timestamp) ────
        sub_header = [""] * num_fixed
        for _ in sorted_timestamps:
            for field in _TIME_FIELDS:
                sub_header.append(_TIME_FIELD_LABELS[field])
        ws.append(sub_header)
        _style_header_row(ws, 2, total_cols, font=_SUBHEADER_FONT, fill=_SUBHEADER_FILL)

        # ── Build a lookup:  (track_id, timestamp) → record ──────────
        rec_lookup: Dict[tuple, ResampledRecord] = {}
        for rec in records:
            rec_lookup[(rec.track_id, rec.timestamp_sec)] = rec

        # ── Data rows (one per track) ────────────────────────────────
        row_count = 0
        for tid in sorted(tracks.keys()):
            recs = tracks[tid]
            first_t = recs[0].timestamp_sec
            last_t = recs[-1].timestamp_sec

            row = [
                tid,
                recs[0].class_name,
                round(first_t, 2),
                round(last_t, 2),
            ]

            for ts in sorted_timestamps:
                rec = rec_lookup.get((tid, ts))
                if rec is not None:
                    row.extend([
                        round(rec.cx_px, 2),
                        round(rec.cy_px, 2),
                        round(rec.speed_px_per_sec, 2),
                        round(rec.cumulative_distance_px, 2),
                    ])
                else:
                    # Track not present at this timestamp
                    row.extend([None, None, None, None])

            ws.append(row)
            row_count += 1

        _auto_column_width(ws, min_width=8, max_width=18)
        _apply_alternating_rows(ws, 3, row_count + 2, total_cols)

        # Freeze the fixed columns + header rows
        ws.freeze_panes = ws.cell(row=3, column=num_fixed + 1)

    def _write_summary_sheet(
        self,
        ws,
        records: List[ResampledRecord],
    ) -> None:
        """Write per-track summary statistics."""
        headers = self._SUMMARY_HEADERS
        ws.append(headers)
        _style_header_row(ws, 1, len(headers))

        # Group records by track_id
        tracks: Dict[int, List[ResampledRecord]] = {}
        for rec in records:
            tracks.setdefault(rec.track_id, []).append(rec)

        row_count = 0
        for tid in sorted(tracks.keys()):
            recs = sorted(tracks[tid], key=lambda r: r.timestamp_sec)
            first_t = recs[0].timestamp_sec
            last_t = recs[-1].timestamp_sec
            duration = last_t - first_t

            total_dist = recs[-1].cumulative_distance_px if recs else 0.0
            speeds = [r.speed_px_per_sec for r in recs]
            avg_speed = sum(speeds) / len(speeds) if speeds else 0.0
            max_speed = max(speeds) if speeds else 0.0
            avg_conf = sum(r.confidence for r in recs) / len(recs) if recs else 0.0

            ws.append([
                tid,
                recs[0].class_id,
                recs[0].class_name,
                round(first_t, 2),
                round(last_t, 2),
                round(duration, 2),
                round(total_dist, 2),
                round(avg_speed, 2),
                round(max_speed, 2),
                len(recs),
                round(avg_conf, 4),
            ])
            row_count += 1

        _auto_column_width(ws)
        _apply_alternating_rows(ws, 2, row_count + 1, len(headers))
        ws.freeze_panes = "A2"

    def _write_metadata_sheet(
        self,
        ws,
        metadata: Dict[str, Any],
    ) -> None:
        """Write processing metadata as key-value pairs."""
        ws.append(["Parameter", "Value"])
        _style_header_row(ws, 1, 2)

        for key, value in metadata.items():
            ws.append([str(key), str(value)])

        _auto_column_width(ws)
